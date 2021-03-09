import openpyxl as op
from aircraft import Operation
from aircraft import Aircraft
from aircraft import Component
class Imp:
    Aircraft_info = {}
    Operations = {}
    Component = {}
    Componento = {}
    def __init__(self,path):
        Workbook = op.load_workbook(path,data_only=True)
        print(Workbook.sheetnames)
        sheet = Workbook["Aircraft Information"]
        print(sheet)
        validsheets = ["Aircraft Information", "Operations", "Component Maintenance Record","Component Maintenance Rec"]
        for sheetname in validsheets:
            if sheetname in Workbook.sheetnames:
                if sheetname == validsheets[0]:
                    self.iAircraft(Workbook[sheetname])
                elif sheetname == validsheets[1]:
                    self.iOperation(Workbook[sheetname])
                elif sheetname == validsheets[2]:
                    self.iComponent(Workbook[sheetname])
                elif sheetname == validsheets[3]:
                    self.iComponento(Workbook[sheetname])
        self.aircraft = Aircraft.Aircraft(self.Aircraft_info,self.Component,self.Componento,self.Operations)

    def iAircraft(self,sheet):
        attr = ["Aircraft Make","Model", "Serial #", "Registration #", "Date made","Hobbs adjust","Tach","Hobbs","Next Insp","Propeller T.T.","Prop.  SMOH","NEXT INSP DUE","Engine T.T.","Eng. SMOH","Time Remaining To Next Insp."]
        for row_cells in sheet.iter_rows():
            for cell in range(len(row_cells)):
                if row_cells[cell].value in attr:
                    self.Aircraft_info[row_cells[cell].value] = row_cells[cell+1].value
        #done

    def iOperation(self,sheet):
        op = {"Inspection":None,"Frequency":None,"Hours":None,"Years":None,"Months":None,"HCW":None,"DCW":None,"C_Tach":None,"NDH":None,"NDD":None,"T_rem":None,"Note":None}
        c = 1
        for row_cells in sheet.iter_rows():
            new = op.copy()
            if 'Operation' in str(row_cells[0].value):
                id = row_cells[0].value
                for cell in range(len(row_cells)):
                    new[list(new.keys())[cell]] = row_cells[cell].value
                self.Operations[id] = Operation.Operation(new)
                c += 1
        #done

    def iComponent(self,sheet):
        comp = {"Item":None, "P_No":None, "S_No":None, "Date":None, "RY":None, "RM":None, "RH":None, "RO":None, "Tot_I":None, "Past_T":None, "Comp_cycle":None, "Date_CW":None, "T_Remove":None,"Sched_Remove":None, "T_RemH":None, "T_remD":None}
        c = 0
        for row_cells in sheet.iter_rows():
            new = comp.copy()
            if str(row_cells[0].value) != "Item" and str(row_cells[0].value) != "None" and c > 2:
                id = row_cells[0].value
                for cell in range(len(row_cells)):
                    try:
                        new[list(new.keys())[cell]] = row_cells[cell].value
                    except:
                        pass
                self.Component[id] = Component.Component(new)
            c += 1
        #done?

    def iComponento(self,sheet):
        compo = {"Item":None, "P_No":None, "S_No":None, "Date":None, "RY":None, "RH":None, "Tot_I":None, "Past_T":None, "Curr_time":None, "AF_Hours":None, "L_Rem":None, "Date_rep":None, "Next_inspH":None, "Next_insp":None, "Type_mjr":None, "Time_mjr":None,"Remarks":None}
        c = 0
        for row_cells in sheet.iter_rows():
            new = compo.copy()

            if str(row_cells[0].value) != "Item" and str(row_cells[0].value) != "None" and c > 2:
                id = row_cells[0].value
                for cell in range(len(row_cells)):
                    try:
                        new[list(new.keys())[cell]] = row_cells[cell].value
                    except:
                        pass
                self.Componento[id] = new
            c += 1

    def get_Aircraft(self):
        return self.aircraft

if __name__ == "__main__":
    a = Imp(r'C:\Users\nateb\PycharmProjects\pyproject\aircraft\N25604_Aircraft_component_list.xlsx')
    print(a.aircraft.get_info("Aircraft Make"))
    print(a.Componento["AIRFRAME"])